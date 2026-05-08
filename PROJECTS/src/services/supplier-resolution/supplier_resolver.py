"""
Supplier Resolution Stage
Sits between classify and scrape in the pipeline.

Entry point: resolve_suppliers(cfg)
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from web_searcher import find_supplier_url
from confidence_scorer import pick_best_url


def _load_master_list(path: str) -> dict:
    """Load master list into {supplier_name: website} dict."""
    df = pd.read_excel(path, usecols=["Supplier Name", "Website"])
    return {
        str(row["Supplier Name"]).strip().upper(): str(row["Website"]).strip()
        for _, row in df.iterrows()
        if row["Supplier Name"] and row["Website"]
    }


def _extract_suppliers(classified_excel: str) -> list:
    """Extract unique supplier names from classified Excel."""
    df = pd.read_excel(classified_excel)
    col = next(
        (c for c in df.columns if "supplier" in c.lower()),
        None
    )
    if not col:
        raise ValueError(
            f"No supplier column found in {classified_excel}. "
            f"Columns: {list(df.columns)}"
        )
    return list(df[col].dropna().str.strip().str.upper().unique())


def _append_to_pending(path: str, rows: list) -> None:
    """Append rows to pending list Excel. Creates file if absent."""
    headers = [
        "Supplier Name", "Suggested URL", "Confidence Score",
        "Search Query", "DuckDuckGo Result", "Bing Result",
        "Status", "Date Added"
    ]
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)

    existing = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    for row in rows:
        if row["Supplier Name"] not in existing:
            ws.append([row.get(h, "") for h in headers])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def _append_to_master_list(path: str, rows: list) -> None:
    """Append high-confidence suppliers to master list."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    existing = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    for row in rows:
        if row["Supplier Name"] not in existing:
            ws.append([row["Supplier Name"], row["Suggested URL"], None])
    wb.save(path)


def _write_resolved_list(path: str, suppliers: list) -> None:
    """Write resolved supplier list for scraper consumption."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Supplier Name", "Website", "Source"])
    for s in suppliers:
        ws.append([s["Supplier Name"], s["Website"], s["Source"]])
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def resolve_suppliers(cfg: dict) -> bool:
    """
    Pipeline stage entry point.

    cfg keys used:
      - classified_excel: path to labeled Excel from classify stage
      - master_list: path to updated_master_list.xlsx
      - supplier_resolution.pending_list_path
      - supplier_resolution.resolved_list_path
      - supplier_resolution.confidence_threshold (default 70)
      - supplier_resolution.search_delay_seconds (default 1.5)
      - supplier_resolution.search_timeout_seconds (default 10)

    Returns True on success, False on failure.
    """
    try:
        res_cfg = cfg.get("supplier_resolution", {})
        master_list_path = cfg["master_list"]
        classified_excel = cfg["classified_excel"]
        pending_path = res_cfg.get(
            "pending_list_path", "data/supplier-pending/new_suppliers_pending.xlsx"
        )
        resolved_path = res_cfg.get(
            "resolved_list_path", "data/supplier-pending/resolved_suppliers.xlsx"
        )
        threshold = res_cfg.get("confidence_threshold", 70)
        delay = res_cfg.get("search_delay_seconds", 1.5)
        timeout = res_cfg.get("search_timeout_seconds", 10)

        print(f"[supplier_resolution] Loading master list: {master_list_path}")
        master = _load_master_list(master_list_path)

        print(f"[supplier_resolution] Extracting suppliers from: {classified_excel}")
        all_suppliers = _extract_suppliers(classified_excel)
        print(f"[supplier_resolution] Found {len(all_suppliers)} unique suppliers")

        known = []
        to_resolve = []
        for name in all_suppliers:
            if name in master:
                known.append({
                    "Supplier Name": name,
                    "Website": master[name],
                    "Source": "master_list"
                })
            else:
                to_resolve.append(name)

        print(f"[supplier_resolution] Known: {len(known)}, Unknown: {len(to_resolve)}")

        high_confidence = []
        low_confidence = []
        now = datetime.now().strftime("%Y-%m-%d")

        for name in to_resolve:
            print(f"[supplier_resolution] Resolving: {name}")
            query = f'"{name}" official website'
            ddg_urls, bing_urls = find_supplier_url(name, delay=delay, timeout=timeout)
            best_url, score = pick_best_url(ddg_urls, bing_urls, name)

            row = {
                "Supplier Name": name,
                "Suggested URL": best_url,
                "Confidence Score": score,
                "Search Query": query,
                "DuckDuckGo Result": ddg_urls[0] if ddg_urls else "",
                "Bing Result": bing_urls[0] if bing_urls else "",
                "Date Added": now,
            }

            if score >= threshold and best_url:
                print(f"  -> High confidence ({score}): {best_url}")
                row["Status"] = "Auto-Added"
                high_confidence.append(row)
            else:
                print(f"  -> Low confidence ({score}): {best_url or 'no result'}")
                row["Status"] = "Pending Review"
                low_confidence.append(row)

        if high_confidence:
            print(f"[supplier_resolution] Adding {len(high_confidence)} to master list")
            _append_to_master_list(master_list_path, high_confidence)

        if low_confidence:
            print(f"[supplier_resolution] Writing {len(low_confidence)} to pending list")
            _append_to_pending(pending_path, low_confidence)

        scrape_list = known + [
            {"Supplier Name": r["Supplier Name"],
             "Website": r["Suggested URL"],
             "Source": "auto_resolved"}
            for r in high_confidence
        ]
        _write_resolved_list(resolved_path, scrape_list)
        print(f"[supplier_resolution] Resolved list written: {len(scrape_list)} suppliers")
        print(f"[supplier_resolution] Done. High: {len(high_confidence)}, "
              f"Low: {len(low_confidence)}, Known: {len(known)}")

        return True

    except Exception as e:
        print(f"[supplier_resolution] ERROR: {e}")
        return False
